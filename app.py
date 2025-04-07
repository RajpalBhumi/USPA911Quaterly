import streamlit as st
import pandas as pd
import pdfplumber
import io
import zipfile
import os
import re
from datetime import date
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

# SECTION I Contact Info (Fixed Contact)
contact_info = {
    "Contact Name": "Seth Tenore",
    "Phone": "877-780-4848",
    "Fax": "506-675-8989",
    "E-mail": "communicationonlinefiling@avalara.com"
}

# Extract basic Section I data dynamically

def extract_basic_data(text):
    lines = text.splitlines()
    data = {}

    for i, line in enumerate(lines):
        l = line.lower()

        if "provider name" in l or "company" in l:
            data["Provider Name"] = lines[i + 1].strip()

        elif "federal tax id" in l or "tax id" in l:
            data["Federal Tax ID"] = lines[i + 1].strip()

        elif "customer id" in l or "pa customer" in l:
            data["Customer ID"] = lines[i + 1].strip()

        elif "address line 1" in l or "street" in l or "road" in l:
            data["Address Line 1"] = lines[i + 1].strip()

        elif "city" in l or "state" in l or "zip" in l:
            data["Address Line 2"] = lines[i + 1].strip()

        elif "filing period" in l:
            data["Period Ending"] = lines[i + 1].strip()

        elif "payment amount" in l:
            data["Payment Amount"] = lines[i + 1].strip()

    return data

# Extract surcharge table values

def extract_surcharge_rows_pdfplumber(file):
    months = {
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    }

    result = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if not row or len(row) < 3:
                        continue
                    row_clean = [str(cell).strip() if cell else "" for cell in row]
                    if row_clean[0] in months:
                        try:
                            assessed = float(row_clean[1].replace(",", "").replace("$", ""))
                            collected = float(row_clean[2].replace(",", "").replace("$", ""))
                            result.append({
                                "month": row_clean[0],
                                "assessed": assessed,
                                "collected": collected
                            })
                        except:
                            continue
    return result

# Fill Excel template with extracted data

def fill_excel_template(template_bytes, data_dict, section_v_data, surcharge_rows):
    wb = load_workbook(filename=template_bytes)
    ws = wb["Remittance Report"]
    ws.protection.sheet = False

    ws["B7"] = data_dict.get("Provider Name", "")
    ws["B8"] = data_dict.get("Federal Tax ID", "")
    ws["B9"] = data_dict.get("Customer ID", "")
    ws["B11"] = data_dict.get("Address Line 1", "")
    ws["B12"] = data_dict.get("Address Line 2", "")
    ws["E7"] = contact_info["Contact Name"]
    ws["E8"] = contact_info["Phone"]
    ws["E9"] = contact_info["Fax"]
    ws["E10"] = contact_info["E-mail"]
    ws["E12"] = data_dict.get("Period Ending", "")

    payment_raw = data_dict.get("Payment Amount", "")
    match = re.search(r"[\d,]+\.\d{2}", payment_raw)
    ws["F13"] = float(match.group(0).replace(",", "")) if match else 0.0

    # ✅ SAFELY UNMERGE and write to Section V
    section_v_ranges = {
        "B41:D41": "B41",
        "E41:F41": "D41",
        "B43:D43": "B43"
    }

    for rng, anchor in section_v_ranges.items():
        if rng in [str(r) for r in ws.merged_cells.ranges]:
            ws.unmerge_cells(rng)

    ws["B41"] = section_v_data["initials"]
    ws["D41"] = section_v_data["title"]
    ws["F41"] = section_v_data["date"]
    ws["B43"] = section_v_data["full_name"]

    start_row_ii = 17
    for i, row in enumerate(surcharge_rows[:3]):
        ws[f"B{start_row_ii + i}"] = row["month"]
        ws[f"C{start_row_ii + i}"] = row["assessed"]
        ws[f"D{start_row_ii + i}"] = row["collected"]

    start_row_iii = 30
    for i, row in enumerate(surcharge_rows[3:6]):
        ws[f"B{start_row_iii + i}"] = row["month"]
        ws[f"C{start_row_iii + i}"] = row["assessed"]
        ws[f"D{start_row_iii + i}"] = row["collected"]

    try:
        logo = ExcelImage("logo.png")
        logo.width = 150
        logo.height = 50
        ws.add_image(logo, "B1")
    except FileNotFoundError:
        print("logo.png not found")

    return wb

# Streamlit UI
st.set_page_config(page_title="911 Remittance Excel Generator", layout="centered")
st.title("📄 Avalara PDF ➝ Branded Excel Report Generator")

st.sidebar.header("✍️ Section V – Certification Info")
initials = st.sidebar.text_input("Initials", "Rhenry")
title = st.sidebar.text_input("Title", "Sr Tax Analyst")
full_name = st.sidebar.text_input("Full Name", "Rachel Henry")
cert_date = st.sidebar.date_input("Date", value=date.today())

section_v_data = {
    "initials": initials,
    "title": title,
    "full_name": full_name,
    "date": cert_date.strftime("%-m/%-d/%Y")
}

template_file = "Template Report.xlsx"
uploaded_files = st.file_uploader("Upload Avalara Confirmation PDF(s)", type="pdf", accept_multiple_files=True)

if uploaded_files:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        for pdf in uploaded_files:
            with st.spinner(f"Processing {pdf.name}..."):
                try:
                    with pdfplumber.open(pdf) as pdf_doc:
                        full_text = "\n".join([page.extract_text() for page in pdf_doc.pages if page.extract_text()])
                    basic_data = extract_basic_data(full_text)
                    surcharge_rows = extract_surcharge_rows_pdfplumber(pdf)

                    with open(template_file, "rb") as f:
                        template_bytes = io.BytesIO(f.read())

                    wb = fill_excel_template(template_bytes, basic_data, section_v_data, surcharge_rows)

                    excel_io = io.BytesIO()
                    wb.save(excel_io)
                    excel_io.seek(0)
                    zipf.writestr(os.path.splitext(pdf.name)[0] + ".xlsx", excel_io.read())
                except Exception as e:
                    st.error(f"Error processing {pdf.name}: {e}")

    zip_buffer.seek(0)
    st.success("✅ All files processed.")
    st.download_button("📦 Download All Reports (ZIP)", data=zip_buffer, file_name="remittance_reports.zip", mime="application/zip")
